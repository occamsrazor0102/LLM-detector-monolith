"""File loaders for xlsx, csv, and pdf input."""

import os
import pandas as pd
from llm_detector.compat import HAS_PYPDF

if HAS_PYPDF:
    from pypdf import PdfReader


def load_xlsx(filepath, sheet=None, prompt_col='prompt', id_col='task_id',
              occ_col='occupation', attempter_col='attempter_name', stage_col='pipeline_stage_name'):
    """Load tasks from an xlsx file. Returns list of dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if sheet:
        ws = wb[sheet]
    else:
        for name in ['FullTaskX', 'Full Task Connected', 'Claim Sheet', 'Sample List']:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        else:
            ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    def find_col(candidates):
        for c in candidates:
            cl = c.lower()
            for i, h in enumerate(headers):
                if cl == h:
                    return i
        for c in candidates:
            cl = c.lower()
            if cl == 'id':
                continue
            for i, h in enumerate(headers):
                if cl in h:
                    return i
        return None

    prompt_idx = find_col([prompt_col, 'prompt', 'text', 'content'])
    id_idx = find_col([id_col, 'task_id', 'id'])
    occ_idx = find_col([occ_col, 'occupation', 'occ'])
    att_idx = find_col([attempter_col, 'attempter', 'claimed_by', 'claimed by'])
    stage_idx = find_col([stage_col, 'stage', 'pipeline_stage'])

    if prompt_idx is None:
        print(f"ERROR: Could not find prompt column. Headers: {headers}")
        return []

    tasks = []
    for row in rows[1:]:
        if not row or len(row) <= prompt_idx:
            continue
        prompt = str(row[prompt_idx]).strip() if row[prompt_idx] else ''
        if len(prompt) < 50:
            continue

        tasks.append({
            'prompt': prompt,
            'task_id': str(row[id_idx])[:20] if id_idx is not None and row[id_idx] else '',
            'occupation': str(row[occ_idx]) if occ_idx is not None and row[occ_idx] else '',
            'attempter': str(row[att_idx]) if att_idx is not None and row[att_idx] else '',
            'stage': str(row[stage_idx]) if stage_idx is not None and row[stage_idx] else '',
        })

    return tasks


def load_csv(filepath, prompt_col='prompt'):
    """Load tasks from CSV."""
    df = pd.read_csv(filepath)
    df = df.fillna('')

    col_map = {c.lower().strip(): c for c in df.columns}

    def resolve_col(*candidates):
        for c in candidates:
            key = c.lower().strip()
            if key in col_map:
                return col_map[key]
        for c in candidates:
            key = c.lower().strip()
            if key == 'id':
                continue
            for mapped_key, actual in col_map.items():
                if key in mapped_key:
                    return actual
        return None

    prompt_actual = resolve_col(prompt_col, 'prompt', 'text', 'content')
    id_actual = resolve_col('task_id', 'id')
    occ_actual = resolve_col('occupation', 'occ')
    att_actual = resolve_col('attempter_name', 'attempter', 'claimed_by')
    stage_actual = resolve_col('pipeline_stage_name', 'stage')

    if prompt_actual is None:
        print(f"ERROR: Could not find prompt column. Columns: {list(df.columns)}")
        return []

    tasks = []
    for _, row in df.iterrows():
        prompt = str(row.get(prompt_actual, ''))
        if len(prompt) < 50:
            continue
        tasks.append({
            'prompt': prompt,
            'task_id': str(row.get(id_actual, ''))[:20] if id_actual else '',
            'occupation': str(row.get(occ_actual, '')) if occ_actual else '',
            'attempter': str(row.get(att_actual, '')) if att_actual else '',
            'stage': str(row.get(stage_actual, '')) if stage_actual else '',
        })
    return tasks


def load_pdf(filepath):
    """Load text from PDF file. Each page becomes a separate task."""
    if not HAS_PYPDF:
        print("ERROR: pypdf not installed. Run: pip install pypdf")
        return []

    reader = PdfReader(filepath)
    tasks = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and len(text.strip()) >= 50:
            tasks.append({
                'prompt': text.strip(),
                'task_id': f"page_{i+1}",
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    if not tasks:
        full_text = '\n'.join(
            page.extract_text() for page in reader.pages
            if page.extract_text()
        ).strip()
        if len(full_text) >= 50:
            tasks.append({
                'prompt': full_text,
                'task_id': 'full_document',
                'occupation': '',
                'attempter': '',
                'stage': '',
            })

    return tasks
